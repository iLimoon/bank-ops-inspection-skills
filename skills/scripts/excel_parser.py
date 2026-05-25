#!/usr/bin/env python3
"""Excel 解析工具 — 从汇总大表解析各 Sheet 的指标数据。

Usage:
    from scripts.excel_parser import parse_summary_xlsx

    data = parse_summary_xlsx('巡检数据汇总_2026W21.xlsx')
    perf = data['1-交易性能']        # 8 条业务线的交易指标
    server = data['2-服务器资源']    # 16 条节点资源指标
    db = data['3-数据库巡检']        # 86 条数据库指标
    dep = data['4-依赖系统']         # 18 条依赖系统记录
    events = data['5-变更与事件']    # 20 条变更事件记录
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.data_utils import safe_float, identify_biz


def norm_header(h):
    """标准化列名：去除换行符和首尾空格。"""
    if h is None:
        return ''
    return str(h).replace('\n', '').replace('\r', '').strip()


def parse_summary_xlsx(fpath, target_week=None):
    """解析汇总大表 Excel，返回所有 Sheet 的结构化数据。

    Args:
        fpath: Excel 文件路径
        target_week: 可选，按周次筛选（如 '2026W21'）

    Returns:
        dict: {sheet_name: [row_dict, ...]}
    """
    import openpyxl
    wb = openpyxl.load_workbook(fpath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # 找表头行（第一个非空行）
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            if row and any(c is not None for c in row):
                # 跳过标题/说明行（通常是合并单元格）
                non_empty = [str(c).strip() for c in row if c is not None]
                if non_empty and len(non_empty) >= 2:
                    header_row = [norm_header(c) for c in row]
                    data_start = i + 1
                    break

        if header_row is None:
            result[sheet_name] = []
            continue

        # 构建列名→索引映射
        col_map = {h: i for i, h in enumerate(header_row) if h}

        # 解析数据行
        data_rows = []
        for row in rows[data_start:]:
            if not row or all(c is None for c in row):
                continue
            values = [c for c in row if c is not None]
            if not values:
                continue

            record = {}
            for col_name, idx in col_map.items():
                if idx < len(row):
                    record[col_name] = row[idx]
                else:
                    record[col_name] = None

            # 周次筛选
            week_cols = [c for c in col_map if '周次' in c]
            if target_week and week_cols:
                week_val = str(record.get(week_cols[0], '')).strip()
                if target_week not in week_val:
                    continue

            data_rows.append(record)

        result[sheet_name] = data_rows

    wb.close()
    return result


def get_biz_rows(sheet_data, biz_name):
    """从 '1-交易性能' sheet 中提取指定业务线的数据。

    Args:
        sheet_data: parse_summary_xlsx() 返回的 sheet 数据
        biz_name: 业务线名称（如 'ESB系统'、'银联无卡'）

    Returns:
        dict 或 None
    """
    biz_cols = ['业务线', '系统名称', '业务系统']
    for row in sheet_data:
        for col in biz_cols:
            if col in row and str(row.get(col, '')).strip() == biz_name:
                return row
    return None


def extract_metric_series(weekly_data, sheet_name, metric_key, biz_filter=None):
    """从多周汇总数据中提取指标的时间序列。

    Args:
        weekly_data: {week_label: parsed_excel_data}
        sheet_name: Sheet 名称
        metric_key: 指标列名
        biz_filter: 可选，按业务线筛选

    Returns:
        list of (week_label, value)
    """
    series = []
    for week in sorted(weekly_data.keys()):
        rows = weekly_data[week].get(sheet_name, [])
        values = []
        for row in rows:
            if biz_filter:
                biz_cols = ['业务线', '系统名称']
                match = any(str(row.get(c, '')).strip() == biz_filter for c in biz_cols)
                if not match:
                    continue
            v = row.get(metric_key)
            if v is not None:
                fv = safe_float(v)
                if fv is not None:
                    values.append(fv)
        if values:
            series.append((week, sum(values) / len(values)))
    return series
